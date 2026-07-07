# coding=utf-8
"""
db_to_excel.py
==============
功能：將 output_db 資料夾下的 .db 檔案轉換為 Excel 報表。
邏輯：
  1. 每個 DB 獨立轉換。
  2. 方向白名單（正面表列）+ 自動去重（向量化）：
     - 先用 OUTPUT_DIRECTIONS 過濾，只保留指定 Direction（IN / OUT / N/A）的紀錄。
     - 有車牌：同車牌時間差 < 5 分鐘視為同一次進場，留 HitCount 最大者。
     - 無車牌(N/A plate)：用較嚴格的 3 秒窗口去重
       （只有極短時間內的 N/A 才視為同一台車的連續幀，相隔較久的視為不同台車各自保留）。
  3. 自動分卷：超過 5000 筆自動拆分。
  4. 多進程預處理圖片：PIL 縮放 + JPEG 壓縮平行處理，大幅加速。

效能設計（因應單一 DB 可能達數千萬筆）：
  - 去重全程 pandas 向量化（無 iterrows），數千萬筆僅需數秒
  - 進度條分母用「去重前總數」，每個 DB 去重後立即補上被刪/被丟筆數，
    進度條讀第一個 DB 即開始動，且能精準跑到 100%

DB schema（events 表，由 state_db.py 產生）：
    id / DeviceCode / CameraCode / TrackID / Plate / Class / ROI / Direction /
    HitCount / VideoTime / CreateTime / ClassImg / PlateImg
    ClassImg / PlateImg 為「相對專案根目錄」的截圖路徑，開圖時補上 BASE_DIR
    save_screenshot=false 時這兩欄為 NULL，Excel 對應格留白
"""

import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PilImage
import io
import os
import sys
import math
import multiprocessing as mp
from functools import partial
from pathlib import Path
from datetime import datetime, timedelta
from tqdm import tqdm

# ==========================================
# 1. 系統配置區
# ==========================================
# 自動偵測腳本所在目錄當作專案根目錄（DB 內截圖路徑是相對於這裡）
BASE_DIR = Path(__file__).resolve().parent

DB_DIR = BASE_DIR / "output_db"          # 輸入：存放 .db 檔案的資料夾
OUTPUT_DIR = BASE_DIR / "output_xlsx"    # 輸出：產生 .xlsx 檔案的資料夾
LIMIT_PER_FILE = 5000                    # 每份 Excel 最大筆數，超過自動分卷
BATCH_SIZE = 200                         # 每個 worker 一次處理的圖片數量

# --- 去重窗口設定 ---
TIME_THRESHOLD_MINUTES = 5               # 有車牌：同車牌在此時間內視為同一次進場（分鐘）
NA_THRESHOLD_SECONDS = 3                 # 無車牌(N/A)：相鄰在此秒數內才視為同一台車連續幀（秒，較嚴格）

# --- 方向輸出白名單（正面表列）---
# 只有 Direction 落在這個清單裡的紀錄才寫進 Excel。
# Direction 欄在 DB 只有三種值：進場、出場、方向未定，對應：
#   "IN"  = 進場   /   "OUT" = 出場   /   "N/A" = 方向未定（位移不足以判定）
# 範例：
#   ["IN", "OUT", "N/A"] → 全部輸出
#   ["OUT", "N/A"]       → 只輸出「出場」與「方向未定」
#   ["IN"]               → 只輸出「進場」
# 註：DB 內方向未定實際存成 "NA"，這裡寫 "N/A" 或 "NA" 都認得（見 _normalize_direction）。
OUTPUT_DIRECTIONS = ["OUT", "N/A"]

# --- 視為「無車牌」的值 ---
# DB 裡辨識失敗會寫成 "N/A"；空字串、None 也一併視為無車牌
_NA_PLATE_VALUES = {"N/A", "", "na", "NA", "None", "none"}


def _normalize_direction(v):
    """把 Direction 值正規化：IN/OUT 照原樣（大寫），其餘（NA/N/A/空/None）一律視為 'N/A'。"""
    s = str(v).strip().upper() if v is not None else ""
    if s == "IN":
        return "IN"
    if s == "OUT":
        return "OUT"
    return "N/A"   # "NA" / "N/A" / "" / None 等都歸為方向未定


# 把使用者設定的白名單也正規化成一組集合，供比對用
_DIRECTION_WHITELIST = {_normalize_direction(x) for x in OUTPUT_DIRECTIONS}


# ==========================================
# 2. 圖片處理函式
# ==========================================
def resize_image_for_excel(image_path, target_width=120):
    """
    讀取單張圖片並縮放為指定寬度，壓縮為 JPEG bytes。

    :param image_path: 圖片檔案完整路徑（已補上 BASE_DIR）
    :param target_width: 縮放後的目標寬度 (px)
    :return: (jpeg_bytes, 縮放後高度) 或 (None, 0) 表示失敗
    """
    if not image_path:
        return None, 0

    path = Path(image_path)
    if not path.exists():
        return None, 0

    try:
        with PilImage.open(path) as img:
            # 依據目標寬度等比例縮放
            ratio = target_width / float(img.size[0])
            new_h = int(img.size[1] * ratio)
            img = img.resize((target_width, new_h), PilImage.LANCZOS).convert("RGB")

            # 壓縮為 JPEG 格式的 bytes（不寫入磁碟）
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=70)
            buf.seek(0)
            return buf.getvalue(), new_h  # 回傳 bytes（可跨進程序列化）
    except Exception:
        return None, 0


def process_image_batch(batch):
    """
    批次處理多張圖片的縮放與壓縮（由子進程呼叫）。
    每個 worker 一次處理一批，減少進程間通訊開銷。

    :param batch: [(df_idx, 車輛截圖完整路徑, 車牌截圖完整路徑), ...]
    :return: {df_idx: {'veh': (bytes, h), 'plate': (bytes, h)}, ...}
    """
    results = {}
    for idx, veh_path, plate_path in batch:
        # 分別處理車輛截圖與車牌截圖
        veh_data = resize_image_for_excel(veh_path) if veh_path else (None, 0)
        plate_data = resize_image_for_excel(plate_path) if plate_path else (None, 0)
        results[idx] = {'veh': veh_data, 'plate': plate_data}
    return results


def preprocess_images_parallel(df, pbar=None):
    """
    多進程平行預處理 DataFrame 中所有圖片。
    這是整個流程最耗時的步驟，透過多核心平行大幅加速。

    DB 內 ClassImg / PlateImg 是相對路徑，這裡補上 BASE_DIR 組成完整路徑再交給子進程。

    :param df: 去重後的資料 DataFrame（需含 ClassImg, PlateImg 欄位）
    :param pbar: 外部傳入的 tqdm 進度條，每處理一批就更新對應筆數
    :return: {df_idx: {'veh': (bytes, h), 'plate': (bytes, h)}, ...} 圖片快取字典
    """
    # 從 DataFrame 收集所有需要處理的圖片路徑（補 BASE_DIR）
    tasks = []
    for idx, row in df.iterrows():
        class_rel = row.get('ClassImg', '') or ''
        plate_rel = row.get('PlateImg', '') or ''
        class_full = str(BASE_DIR / class_rel) if class_rel else ''
        plate_full = str(BASE_DIR / plate_rel) if plate_rel else ''
        tasks.append((idx, class_full, plate_full))

    # 將任務切分成多個批次，每批 BATCH_SIZE 筆
    batches = [tasks[i:i + BATCH_SIZE] for i in range(0, len(tasks), BATCH_SIZE)]

    # worker 數量不超過 CPU 核心數，也不超過批次數
    num_workers = min(mp.cpu_count(), len(batches)) if batches else 1
    all_results = {}

    with mp.Pool(processes=num_workers) as pool:
        # imap_unordered：不保證順序但更快，結果透過 idx 對應回 DataFrame
        for batch_result in pool.imap_unordered(process_image_batch, batches):
            all_results.update(batch_result)
            # 更新外部進度條（以筆數為單位）
            if pbar is not None:
                pbar.update(len(batch_result))

    return all_results


# ==========================================
# 3. 去重邏輯（向量化，車牌 / N/A 分流）
# ==========================================
def _dedup_by_time_group(df, threshold):
    """
    對「同一辨識群組」的 DataFrame 做時間分組去重（向量化核心）。

    呼叫前 df 必須已按「分組鍵 + 時間」排序，且帶有 _grpkey 欄（分組鍵）與 temp_dt 欄。
    規則：同分組鍵內，相鄰時間差 > threshold → 切新組；每組留 HitCount 最大那筆。

    :param df: 已排序、含 _grpkey 與 temp_dt 的 DataFrame
    :param threshold: pd.Timedelta，分組時間窗口
    :return: 去重後的 DataFrame（保留原欄位，不含暫存欄）
    """
    if df.empty:
        return df

    same_key = df['_grpkey'].eq(df['_grpkey'].shift())     # 與上一筆同分組鍵
    time_gap = df['temp_dt'].diff()                        # 與上一筆時間差
    within = same_key & (time_gap <= threshold)            # 同鍵且時間差在窗口內 → 同一次
    df = df.copy()
    df['_grp'] = (~within).cumsum()                        # 切組

    keep_idx = df.groupby('_grp')['HitCount'].idxmax()     # 每組取 HitCount 最大
    return df.loc[keep_idx]


def filter_duplicates(df, db_name="", pbar=None):
    """
    對 DataFrame 去重（pandas 向量化），車牌與 N/A 分流處理。

    流程：
      0. 方向白名單過濾：只留 Direction 在 OUTPUT_DIRECTIONS 內的紀錄（正面表列）
      1. 把剩下的資料分成「有車牌」與「無車牌(N/A)」兩群
      2. 有車牌群：用 TIME_THRESHOLD_MINUTES（5 分鐘）做時間分組去重
      3. 無車牌群：用 NA_THRESHOLD_SECONDS（3 秒，較嚴格）去重
         （只有極短時間內的 N/A 才當同一台車的連續幀，相隔較久者各自保留）
      4. 兩群合併，依 id 排序

    :param df: 原始 DataFrame（需含 Plate, VideoTime, HitCount, id 欄位）
    :param db_name: DB 名稱（用於 log 顯示）
    :param pbar: 外部進度條；去重完成後 update(被刪+被丟筆數) 把進度補上
    :return: 去重後的 DataFrame
    """
    if df.empty:
        return df

    before_count = len(df)

    # --- 方向白名單過濾（正面表列）：只留 Direction 在 OUTPUT_DIRECTIONS 內的紀錄 ---
    df = df.copy()
    dir_norm = df['Direction'].map(_normalize_direction)
    df = df[dir_norm.isin(_DIRECTION_WHITELIST)]
    dir_removed = before_count - len(df)   # 因方向白名單被排除的筆數

    if df.empty:
        # 全被方向白名單擋掉：把整批補進進度條後直接回空表
        if pbar is not None and before_count > 0:
            pbar.update(before_count)
        tqdm.write(f"  [{db_name}] 方向白名單 {OUTPUT_DIRECTIONS}：{before_count} 筆全數排除")
        return df.iloc[0:0].drop(columns=['temp_dt'], errors='ignore') if 'temp_dt' in df.columns else df

    # 時間轉換
    df['temp_dt'] = pd.to_datetime(df['VideoTime'], format='%H:%M:%S')

    # 分流：有車牌 vs 無車牌(N/A / 空)
    plate_norm = df['Plate'].fillna("").astype(str).str.strip()
    is_na = plate_norm.isin(_NA_PLATE_VALUES)

    df_plate = df[~is_na]    # 有車牌
    df_na = df[is_na]        # 無車牌

    result_parts = []

    # --- 有車牌：用車牌當分組鍵，5 分鐘窗口 ---
    if not df_plate.empty:
        d = df_plate.copy()
        d['_grpkey'] = d['Plate'].astype(str)
        d = d.sort_values(by=['_grpkey', 'temp_dt']).reset_index(drop=True)
        kept = _dedup_by_time_group(d, pd.Timedelta(minutes=TIME_THRESHOLD_MINUTES))
        result_parts.append(kept)

    # --- 無車牌(N/A)：用較嚴格的 3 秒窗口去重（是否輸出已由方向白名單決定）---
    na_kept_count = 0
    if not df_na.empty:
        d = df_na.copy()
        # N/A 全部同一個分組鍵（用 CameraCode 區隔不同 cam，避免跨 cam 誤併；
        # 同一 DB 通常同一個 CameraCode，這裡仍以 CameraCode 為鍵較嚴謹）
        d['_grpkey'] = d['CameraCode'].astype(str)
        d = d.sort_values(by=['_grpkey', 'temp_dt']).reset_index(drop=True)
        kept = _dedup_by_time_group(d, pd.Timedelta(seconds=NA_THRESHOLD_SECONDS))
        result_parts.append(kept)
        na_kept_count = len(kept)

    # 合併兩群，清暫存欄，依 id 排序
    if result_parts:
        df_clean = pd.concat(result_parts, ignore_index=True)
        df_clean = df_clean.drop(columns=['temp_dt', '_grpkey', '_grp'], errors='ignore')
        df_clean = df_clean.sort_values(by='id').reset_index(drop=True)
    else:
        df_clean = df.iloc[0:0].drop(columns=['temp_dt'], errors='ignore')

    removed = before_count - len(df_clean)

    # log：顯示方向白名單排除數 + 無車牌去重結果
    na_total = len(df_na)
    tqdm.write(f"  [{db_name}] 去重: {before_count} → {len(df_clean)} 筆 "
               f"(共減少 {removed} 筆；方向白名單排除 {dir_removed} 筆、無車牌 {na_total}→{na_kept_count} 筆)")

    # 把「被刪+被丟」的筆數補進進度條（分母是去重前總數，這些筆不會經過預處理 update）
    if pbar is not None and removed > 0:
        pbar.update(removed)

    return df_clean


# ==========================================
# 4. Excel 寫入
# ==========================================
def write_dataframe_to_excel(df, output_path, image_cache, pbar=None):
    """
    將 DataFrame 寫入 Excel 檔案，圖片從預處理快取中取用。
    openpyxl 不是 thread-safe，因此寫入必須在單一執行緒中進行。
    但因為圖片已經是壓縮好的 bytes，插入速度非常快。

    :param df: 要寫入的資料 DataFrame
    :param output_path: 輸出的 .xlsx 檔案路徑
    :param image_cache: 圖片快取字典 {idx: {'veh': (bytes, h), 'plate': (bytes, h)}}
    :param pbar: 外部傳入的 tqdm 進度條（可選）
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "車輛辨識結果"

    # 寫入標頭列
    headers = ["ID", "車輛截圖", "車種", "車牌截圖", "車號", "ROI", "方向", "次數", "時間軸", "時間點"]
    ws.append(headers)

    # 設定各欄位寬度（讓圖片和文字顯示合理）
    ws.column_dimensions["B"].width = 18   # 車輛截圖欄
    ws.column_dimensions["D"].width = 18   # 車牌截圖欄
    ws.column_dimensions["I"].width = 15   # 時間軸欄
    ws.column_dimensions["J"].width = 20   # 時間點欄

    # 逐筆寫入資料與圖片
    for idx, row in df.iterrows():
        excel_row = idx + 2   # Excel 第 1 列是標頭，資料從第 2 列開始

        # 寫入文字欄位
        ws.cell(excel_row, 1, row['TrackID'])
        ws.cell(excel_row, 3, row['Class'])
        ws.cell(excel_row, 5, row['Plate'])
        ws.cell(excel_row, 6, row['ROI'])
        ws.cell(excel_row, 7, row['Direction'])
        ws.cell(excel_row, 8, row['HitCount'])
        ws.cell(excel_row, 9, row['VideoTime'])
        ws.cell(excel_row, 10, row['CreateTime'])

        # 從快取取出預處理好的圖片 bytes
        cached = image_cache.get(idx, {})

        # 插入車輛截圖（B 欄）
        veh_bytes, veh_h = cached.get('veh', (None, 0))
        if veh_bytes:
            img = XLImage(io.BytesIO(veh_bytes))
            img.anchor = f"B{excel_row}"
            ws.add_image(img)
            ws.row_dimensions[excel_row].height = veh_h * 0.75  # 像素轉換為 Excel 列高

        # 插入車牌截圖（D 欄）
        plate_bytes, plate_h = cached.get('plate', (None, 0))
        if plate_bytes:
            img = XLImage(io.BytesIO(plate_bytes))
            img.anchor = f"D{excel_row}"
            ws.add_image(img)
            # 取兩張圖片中較高的那個作為列高
            current_h = ws.row_dimensions[excel_row].height or 0
            if plate_h * 0.75 > current_h:
                ws.row_dimensions[excel_row].height = plate_h * 0.75

    # 儲存 Excel 檔案
    wb.save(output_path)


# ==========================================
# 5. 單一 DB 處理流程
# ==========================================
def process_single_db(db_path, pbar):
    """
    處理單一 .db 檔案的完整流程：
    讀取 DB → 向量化去重（車牌/N/A 分流）→ 多進程預處理圖片 → 寫入 Excel（必要時自動分卷）

    :param db_path: .db 檔案的 Path 物件
    :param pbar: 外部總進度條（以筆數為單位，分母為去重前總數）
    """
    # 讀取資料庫（events 表）
    try:
        conn = sqlite3.connect(db_path)
        df = pd.read_sql_query("SELECT * FROM events ORDER BY id", conn)
        conn.close()
    except Exception as e:
        tqdm.write(f"\n[錯誤] 無法讀取 DB {db_path}: {e}")
        return

    if df.empty:
        return

    # 步驟 1：向量化去重（車牌 5 分鐘、N/A 3 秒或排除；去重後把刪除筆數補進進度條）
    df = filter_duplicates(df, db_name=db_path.stem, pbar=pbar)
    total_records = len(df)

    if total_records == 0:
        return

    # 步驟 2：多進程平行預處理所有圖片（最耗時的部分，逐批 update 進度條）
    pbar.set_description(f"{db_path.stem} | 預處理圖片")
    image_cache = preprocess_images_parallel(df, pbar)

    # 步驟 3：寫入 Excel（圖片已是 bytes 快取，插入很快）
    base_name = db_path.stem

    if total_records <= LIMIT_PER_FILE:
        # 不需分卷：直接寫成一份 Excel
        out_path = OUTPUT_DIR / f"{base_name}_output.xlsx"
        pbar.set_description(f"{db_path.stem} | 寫入 Excel")
        write_dataframe_to_excel(df, out_path, image_cache)
    else:
        # 需要分卷：依照 LIMIT_PER_FILE 切分
        total_parts = math.ceil(total_records / LIMIT_PER_FILE)

        for i in range(total_parts):
            start_idx = i * LIMIT_PER_FILE
            end_idx = min((i + 1) * LIMIT_PER_FILE, total_records)

            # 切出該分卷的資料子集，重設索引從 0 開始
            df_part = df.iloc[start_idx:end_idx].reset_index(drop=True)

            # 重新映射 image_cache 的 key（因為 reset_index 後 idx 從 0 開始）
            part_cache = {}
            for new_idx, old_idx in enumerate(range(start_idx, end_idx)):
                if old_idx in image_cache:
                    part_cache[new_idx] = image_cache[old_idx]

            out_path = OUTPUT_DIR / f"{base_name}_output_part{i + 1}.xlsx"
            pbar.set_description(f"{db_path.stem} | 寫入 Part {i + 1}/{total_parts}")
            write_dataframe_to_excel(df_part, out_path, part_cache)


# ==========================================
# 6. 主程式進入點
# ==========================================
def main():
    """
    主函式：掃描所有 .db 檔案，依序處理。

    進度條設計：
      分母用「去重前總筆數」（開頭快速 SELECT COUNT(*) 算出），
      進度條讀第一個 DB 即開始動；每個 DB 去重後補上被刪/被丟筆數，
      預處理階段逐批 update，最終精準跑到 100%。
    """
    if not DB_DIR.exists():
        print(f"[錯誤] 找不到資料夾: {DB_DIR}")
        return

    db_files = list(DB_DIR.glob("*.db"))
    if not db_files:
        print(f"[提示] 在 {DB_DIR} 找不到任何 .db 檔案。")
        return

    # 方向白名單提示
    print(f"[INFO] 方向白名單（只輸出這些 Direction）：{OUTPUT_DIRECTIONS}")

    # 預先掃描所有 DB 的筆數（僅 SELECT COUNT，非常快），當進度條分母
    total_all = 0
    for db_path in db_files:
        try:
            conn = sqlite3.connect(db_path)
            count = pd.read_sql_query(
                "SELECT COUNT(*) as cnt FROM events", conn
            ).iloc[0]['cnt']
            conn.close()
            total_all += count
        except Exception:
            pass

    print(f"[INFO] {len(db_files)} 個資料庫，共約 {total_all} 筆")

    # 建立單一總進度條，以「筆」為單位貫穿全程
    # 進度組成 = 各 DB(去重刪除/排除筆數 + 去重後預處理筆數)，加總正好 = total_all → 100%
    with tqdm(total=total_all, unit="筆", dynamic_ncols=True) as pbar:
        for db_path in db_files:
            process_single_db(db_path, pbar)

    print(f"✅ 全部完成！請查看 {OUTPUT_DIR} 資料夾。")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n🛑 中斷")
        sys.exit(0)
